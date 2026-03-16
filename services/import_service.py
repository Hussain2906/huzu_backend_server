from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any
from uuid import uuid4

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.db.models import (
    Customer,
    InventoryLedger,
    Invoice,
    InvoiceType,
    PaymentMode,
    Product,
    ProductCategory,
    StockItem,
    StockReason,
    Supplier,
    TaxMode,
    User,
)
from app.services.invoice_service import create_invoice


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_bool(value: str | None, default: bool | None = None) -> bool | None:
    raw = _clean(value).lower()
    if not raw:
        return default
    if raw in ("true", "yes", "y", "1"):
        return True
    if raw in ("false", "no", "n", "0"):
        return False
    return default


def _parse_float(value: str | None) -> float | None:
    raw = _clean(value)
    if raw == "":
        return None
    try:
        return float(raw)
    except Exception:
        return None


def _parse_date(value: str | None) -> datetime | None:
    raw = _clean(value)
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw)
    except Exception:
        pass
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt)
        except Exception:
            continue
    return None


def _clean_date_string(value: str | None) -> str | None:
    raw = _clean(value)
    if not raw:
        return None
    parsed = _parse_date(value)
    if not parsed:
        return None
    return parsed.date().isoformat()


def parse_tabular_upload(file: UploadFile) -> tuple[list[dict[str, str]], list[str]]:
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File name is required")
    filename = file.filename.lower()
    content = file.file.read()
    if filename.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except Exception:
            text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return [], []
        rows = []
        for row in reader:
            cleaned = {(_clean(k).lower()): _clean(v) for k, v in row.items() if k is not None}
            rows.append(cleaned)
        return rows, [(_clean(h).lower()) for h in reader.fieldnames]
    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet = workbook.active
        data = list(sheet.iter_rows(values_only=True))
        if not data:
            return [], []
        headers = [(_clean(cell).lower()) for cell in (data[0] or [])]
        if not headers or not any(headers):
            return [], []
        rows: list[dict[str, str]] = []
        for row in data[1:]:
            cleaned = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = row[index] if row and index < len(row) else ""
                cleaned[header] = _clean(value)
            if any(value for value in cleaned.values()):
                rows.append(cleaned)
        return rows, headers
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only CSV and XLSX files are supported")


def parse_tabular_upload_content(filename: str, content: bytes) -> tuple[list[dict[str, str]], list[str]]:
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except Exception:
            text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return [], []
        rows = []
        for row in reader:
            cleaned = {(_clean(k).lower()): _clean(v) for k, v in row.items() if k is not None}
            rows.append(cleaned)
        return rows, [(_clean(h).lower()) for h in reader.fieldnames]
    if lowered.endswith(".xlsx") or lowered.endswith(".xlsm"):
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet = workbook.active
        data = list(sheet.iter_rows(values_only=True))
        if not data:
            return [], []
        headers = [(_clean(cell).lower()) for cell in (data[0] or [])]
        if not headers or not any(headers):
            return [], []
        rows: list[dict[str, str]] = []
        for row in data[1:]:
            cleaned = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = row[index] if row and index < len(row) else ""
                cleaned[header] = _clean(value)
            if any(value for value in cleaned.values()):
                rows.append(cleaned)
        return rows, headers
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only CSV and XLSX files are supported")


def _normalize_header(value: Any) -> str:
    cleaned = _clean(value).lower()
    if not cleaned:
        return ""
    for char in ("-", "/", "\\", "(", ")", ".", ":"):
        cleaned = cleaned.replace(char, " ")
    return " ".join(cleaned.split())


def _generate_product_code(db: Session, user: User) -> str:
    while True:
        code = f"PRD-{uuid4().hex[:8].upper()}"
        exists = (
            db.query(Product)
            .filter(Product.company_id == user.company_id, Product.product_code == code)
            .first()
        )
        if not exists:
            return code


def parse_stock_summary_report_content(filename: str, content: bytes) -> tuple[list[dict[str, str]], list[str]]:
    lowered = filename.lower()
    if not (lowered.endswith(".xlsx") or lowered.endswith(".xlsm")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Stock summary report must be XLSX/XLSM")

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    expected = {
        "sr no": "sr_no",
        "item name": "product_name",
        "item category": "category",
        "current stock": "qty_on_hand",
        "minimum stock": "reorder_level",
        "sale price": "selling_rate",
        "purchase price": "purchase_rate",
        "stock value": "stock_value",
        "stock valuation": "stock_valuation",
    }
    required = {"item name", "current stock"}
    preferred_sheets = sorted(
        workbook.worksheets,
        key=lambda ws: (0 if ("stock" in ws.title.lower() and "summary" in ws.title.lower()) else 1, ws.title.lower()),
    )
    for sheet in preferred_sheets:
        rows = sheet.iter_rows(values_only=True)
        header_map: dict[int, str] | None = None
        headers: list[str] = []
        data_rows: list[dict[str, str]] = []
        for row in rows:
            normalized = [_normalize_header(value) for value in row]
            present = {value for value in normalized if value}
            if header_map is None:
                if required.issubset(present) and len(present.intersection(expected.keys())) >= 4:
                    header_map = {
                        index: expected[value]
                        for index, value in enumerate(normalized)
                        if value in expected
                    }
                    headers = list(dict.fromkeys(header_map.values()))
                continue

            if not header_map:
                continue
            cleaned_row = {
                header_map[index]: _clean(row[index] if index < len(row) else "")
                for index in header_map
            }
            if not any(cleaned_row.values()):
                continue
            if not _clean(cleaned_row.get("product_name")) and not _clean(cleaned_row.get("qty_on_hand")):
                continue
            data_rows.append(cleaned_row)
        if header_map and data_rows:
            return data_rows, headers

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Could not find stock summary table headers in workbook",
    )


def _template_definition(module: str) -> dict[str, Any]:
    templates = {
        "products": {
            "headers": [
                "name",
                "product_code",
                "category",
                "selling_rate",
                "unit",
                "mrp",
                "purchase_rate",
                "taxable",
                "tax_rate",
                "price_tax_mode",
                "cess_rate",
                "hsn",
                "barcode",
                "description",
                "item_type",
                "opening_stock",
                "low_stock_alert",
                "storage_location",
                "bulk_purchase_unit",
                "retail_per_bulk_ratio",
                "expiry_date",
                "show_online",
                "online_sell_price",
            ],
            "sample": {
                "name": "Sample Product",
                "product_code": "PRD-001",
                "category": "General",
                "selling_rate": "100",
                "unit": "Nos",
                "mrp": "120",
                "purchase_rate": "80",
                "taxable": "TRUE",
                "tax_rate": "18",
                "price_tax_mode": "WITHOUT_TAX",
                "cess_rate": "0",
                "hsn": "1234",
                "barcode": "890100000001",
                "description": "Sample retail product",
                "item_type": "PRODUCT",
                "opening_stock": "50",
                "low_stock_alert": "5",
                "storage_location": "Rack A1",
                "bulk_purchase_unit": "Box",
                "retail_per_bulk_ratio": "10",
                "expiry_date": "2027-12-31",
                "show_online": "FALSE",
                "online_sell_price": "110",
            },
        },
        "purchases": {
            "headers": [
                "invoice_no",
                "invoice_date",
                "tax_mode",
                "is_interstate",
                "supplier_name",
                "supplier_phone",
                "supplier_gstin",
                "supplier_address",
                "payment_status",
                "payment_mode",
                "payment_reference",
                "round_off",
                "product_name",
                "product_code",
                "category",
                "description",
                "hsn",
                "qty",
                "unit",
                "price",
                "discount_percent",
                "taxable",
                "tax_rate",
            ],
            "sample": {
                "invoice_no": "P-0001",
                "invoice_date": "2026-02-05",
                "tax_mode": "GST",
                "is_interstate": "FALSE",
                "supplier_name": "Sample Supplier",
                "supplier_phone": "9000000000",
                "supplier_gstin": "22AAAAA0000A1Z5",
                "supplier_address": "Market Road",
                "payment_status": "PAID",
                "payment_mode": "CASH",
                "payment_reference": "Paid in cash",
                "round_off": "0",
                "product_name": "Sample Product",
                "product_code": "PRD-001",
                "category": "General",
                "description": "Sample Product",
                "hsn": "1234",
                "qty": "10",
                "unit": "Nos",
                "price": "80",
                "discount_percent": "0",
                "taxable": "TRUE",
                "tax_rate": "18",
            },
        },
        "sales": {
            "headers": [
                "invoice_no",
                "invoice_date",
                "tax_mode",
                "is_interstate",
                "customer_name",
                "customer_phone",
                "customer_gstin",
                "customer_address",
                "payment_mode",
                "payment_reference",
                "round_off",
                "product_name",
                "product_code",
                "category",
                "description",
                "hsn",
                "qty",
                "unit",
                "price",
                "discount_percent",
                "taxable",
                "tax_rate",
            ],
            "sample": {
                "invoice_no": "S-0001",
                "invoice_date": "2026-02-05",
                "tax_mode": "GST",
                "is_interstate": "FALSE",
                "customer_name": "Sample Customer",
                "customer_phone": "9000000000",
                "customer_gstin": "22AAAAA0000A1Z5",
                "customer_address": "Main Street",
                "payment_mode": "UPI",
                "payment_reference": "UPI ref 1234",
                "round_off": "0",
                "product_name": "Sample Product",
                "product_code": "PRD-001",
                "category": "General",
                "description": "Sample Product",
                "hsn": "1234",
                "qty": "2",
                "unit": "Nos",
                "price": "100",
                "discount_percent": "0",
                "taxable": "TRUE",
                "tax_rate": "18",
            },
        },
        "inventory": {
            "headers": [
                "product_name",
                "product_code",
                "category",
                "hsn",
                "unit",
                "qty_on_hand",
                "selling_rate",
                "purchase_rate",
                "taxable",
                "tax_rate",
            ],
            "sample": {
                "product_name": "Sample Product",
                "product_code": "PRD-001",
                "category": "General",
                "hsn": "1234",
                "unit": "Nos",
                "qty_on_hand": "100",
                "selling_rate": "100",
                "purchase_rate": "80",
                "taxable": "TRUE",
                "tax_rate": "18",
            },
        },
    }

    template = templates.get(module)
    if not template:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown template type")
    return template


def build_template(module: str, output_format: str = "csv") -> tuple[bytes, str, str]:
    template = _template_definition(module)
    normalized_format = (output_format or "csv").lower()
    if normalized_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = module.title()
        sheet.append(template["headers"])
        sheet.append([template["sample"].get(header, "") for header in template["headers"]])
        stream = io.BytesIO()
        workbook.save(stream)
        return (
            stream.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"{module}_template.xlsx",
        )

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=template["headers"])
    writer.writeheader()
    writer.writerow(template["sample"])
    return output.getvalue().encode("utf-8"), "text/csv", f"{module}_template.csv"


def _ensure_category(db: Session, user: User, name: str | None) -> str | None:
    category_id, _ = _ensure_category_result(db, user, name)
    return category_id


def _ensure_category_result(db: Session, user: User, name: str | None) -> tuple[str | None, bool]:
    if not name:
        return None, False
    trimmed = name.strip()
    if not trimmed:
        return None, False
    existing = (
        db.query(ProductCategory)
        .filter(ProductCategory.company_id == user.company_id, func.lower(ProductCategory.name) == trimmed.lower())
        .first()
    )
    if existing:
        return existing.id, False
    category = ProductCategory(company_id=user.company_id, name=trimmed)
    db.add(category)
    db.flush()
    return category.id, True


def _find_category_id(db: Session, user: User, name: str | None) -> str | None:
    trimmed = _clean(name)
    if not trimmed:
        return None
    existing = (
        db.query(ProductCategory)
        .filter(ProductCategory.company_id == user.company_id, func.lower(ProductCategory.name) == trimmed.lower())
        .first()
    )
    return existing.id if existing else None


def _product_extra_data(payload: dict[str, Any], existing: dict[str, Any] | None = None) -> dict[str, Any]:
    extra = dict(existing or {})
    numeric_fields = {
        "mrp": "mrp",
        "online_sell_price": "online_sell_price",
        "cess_rate": "cess_rate",
        "retail_per_bulk_ratio": "retail_per_bulk_ratio",
    }
    text_fields = {
        "price_tax_mode": "price_tax_mode",
        "barcode": "barcode",
        "description": "description",
        "item_type": "item_type",
        "storage_location": "storage_location",
        "bulk_purchase_unit": "bulk_purchase_unit",
    }
    for source_key, target_key in numeric_fields.items():
        value = _parse_float(payload.get(source_key))
        if value is not None:
            extra[target_key] = value
    for source_key, target_key in text_fields.items():
        value = _clean(payload.get(source_key))
        if value:
            extra[target_key] = value
    show_online = _parse_bool(payload.get("show_online"), None)
    if show_online is not None:
        extra["show_online"] = show_online
    expiry_date = _clean_date_string(payload.get("expiry_date"))
    if expiry_date:
        extra["expiry_date"] = expiry_date
    return extra


def _upsert_stock_qty(
    db: Session,
    user: User,
    product: Product,
    qty: float | None,
    *,
    notes: str,
) -> None:
    if qty is None:
        return
    stock = (
        db.query(StockItem)
        .filter(StockItem.company_id == user.company_id, StockItem.product_id == product.id)
        .first()
    )
    current_qty = float(stock.qty_on_hand) if stock and stock.qty_on_hand is not None else 0.0
    target_qty = float(qty)
    if stock is None and target_qty == 0:
        return
    if stock is None:
        stock = StockItem(company_id=user.company_id, product_id=product.id, qty_on_hand=0)
        db.add(stock)
        db.flush()
    delta = target_qty - current_qty
    stock.qty_on_hand = target_qty
    stock.updated_at = datetime.utcnow()
    if delta != 0:
        db.add(
            InventoryLedger(
                company_id=user.company_id,
                product_id=product.id,
                qty_change=delta,
                reason=StockReason.ADJUSTMENT,
                ref_type="import",
                ref_id=None,
                notes=notes,
                created_by=user.id,
            )
        )


def _resolve_product(
    db: Session,
    user: User,
    product_code: str | None,
    product_name: str | None,
    payload: dict[str, Any],
    errors: list[dict],
    row_num: int,
    *,
    options: dict[str, Any] | None = None,
    stats: dict[str, int] | None = None,
) -> Product | None:
    options = options or {}
    code = (product_code or "").strip()
    name = (product_name or "").strip()
    product = None
    if code:
        product = (
            db.query(Product)
            .filter(Product.company_id == user.company_id, Product.product_code == code)
            .first()
        )
    if not product and name:
        product = (
            db.query(Product)
            .filter(Product.company_id == user.company_id, func.lower(Product.name) == name.lower())
            .first()
        )
    if product:
        if stats is not None:
            stats["matched_existing_products"] = stats.get("matched_existing_products", 0) + 1
        if options.get("create_missing_categories", True):
            category_id, category_created = _ensure_category_result(db, user, payload.get("category"))
        else:
            category_id, category_created = _find_category_id(db, user, payload.get("category")), False
        if category_created and stats is not None:
            stats["categories_created"] = stats.get("categories_created", 0) + 1
        if category_id and (not product.category_id or options.get("update_existing_category", True)):
            product.category_id = category_id
        if payload.get("hsn"):
            product.hsn = payload.get("hsn") or product.hsn
        if payload.get("unit"):
            product.unit = payload.get("unit") or product.unit
        parsed_selling = _parse_float(payload.get("selling_rate"))
        parsed_purchase = _parse_float(payload.get("purchase_rate"))
        parsed_tax = _parse_float(payload.get("tax_rate"))
        parsed_taxable = _parse_bool(payload.get("taxable"), None)
        parsed_reorder_level = _parse_float(payload.get("low_stock_alert") or payload.get("reorder_level"))
        update_existing_prices = bool(options.get("update_existing_prices", True))
        update_existing_reorder = bool(options.get("update_existing_reorder_level", True))
        if parsed_selling is not None and (update_existing_prices or product.selling_rate is None):
            product.selling_rate = parsed_selling
        if parsed_purchase is not None and (update_existing_prices or product.purchase_rate is None):
            product.purchase_rate = parsed_purchase
        if parsed_tax is not None:
            product.tax_rate = parsed_tax
        if parsed_taxable is not None:
            product.taxable = parsed_taxable
        if parsed_reorder_level is not None and (update_existing_reorder or product.reorder_level is None):
            product.reorder_level = parsed_reorder_level
        product.extra_json = _product_extra_data(payload, product.extra_json or {})
        return product

    if not name:
        errors.append({"row": row_num, "field": "product_name", "message": "Product name required to create product"})
        return None

    if not options.get("create_missing_products", True):
        errors.append({"row": row_num, "field": "product_name", "message": "Product not found and create_missing_products is disabled"})
        return None

    category_id = None
    category_created = False
    category_name = payload.get("category")
    if category_name:
        if options.get("create_missing_categories", True):
            category_id, category_created = _ensure_category_result(db, user, category_name)
        else:
            category_id = _find_category_id(db, user, category_name)
    elif options.get("fallback_category_name"):
        if options.get("create_missing_categories", True):
            category_id, category_created = _ensure_category_result(db, user, options["fallback_category_name"])
        else:
            category_id = _find_category_id(db, user, options["fallback_category_name"])
    if category_created and stats is not None:
        stats["categories_created"] = stats.get("categories_created", 0) + 1
    taxable = _parse_bool(payload.get("taxable"), True)
    tax_rate = _parse_float(payload.get("tax_rate"))
    selling_rate = _parse_float(payload.get("selling_rate"))
    purchase_rate = _parse_float(payload.get("purchase_rate"))
    reorder_level = _parse_float(payload.get("low_stock_alert") or payload.get("reorder_level"))
    hsn = payload.get("hsn") or None
    unit = payload.get("unit") or None
    code_value = code or _generate_product_code(db, user)

    if code_value:
        existing = (
            db.query(Product)
            .filter(Product.company_id == user.company_id, Product.product_code == code_value)
            .first()
        )
        if existing:
            errors.append({"row": row_num, "field": "product_code", "message": "Product code already exists"})
            return None

    product = Product(
        company_id=user.company_id,
        name=name,
        product_code=code_value,
        category_id=category_id,
        hsn=hsn,
        selling_rate=selling_rate,
        purchase_rate=purchase_rate,
        unit=unit,
        taxable=bool(taxable) if taxable is not None else True,
        tax_rate=tax_rate,
        reorder_level=reorder_level,
        extra_json=_product_extra_data(payload, {}),
    )
    db.add(product)
    db.flush()
    if stats is not None:
        stats["products_created"] = stats.get("products_created", 0) + 1
    return product


def import_products(db: Session, user: User, rows: list[dict[str, str]], headers: list[str]) -> dict:
    required = {"name"}
    if not required.issubset(set(headers)):
        missing = ", ".join(sorted(required - set(headers)))
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Missing required columns: {missing}")

    errors: list[dict] = []
    seen_codes: set[str] = set()

    for idx, row in enumerate(rows):
        row_num = idx + 2
        name = row.get("name", "").strip()
        if not name:
            errors.append({"row": row_num, "field": "name", "message": "Name required"})
            continue
        code = row.get("product_code", "").strip()
        if code:
            key = code.lower()
            if key in seen_codes:
                errors.append({"row": row_num, "field": "product_code", "message": "Duplicate product code in file"})
            seen_codes.add(key)
        for field in (
            "selling_rate",
            "purchase_rate",
            "tax_rate",
            "mrp",
            "online_sell_price",
            "cess_rate",
            "opening_stock",
            "low_stock_alert",
            "retail_per_bulk_ratio",
        ):
            raw = row.get(field)
            if _clean(raw) and _parse_float(raw) is None:
                errors.append({"row": row_num, "field": field, "message": "Invalid number"})
        taxable_raw = row.get("taxable")
        if _clean(taxable_raw) and _parse_bool(taxable_raw, None) is None:
            errors.append({"row": row_num, "field": "taxable", "message": "Invalid boolean"})
        show_online_raw = row.get("show_online")
        if _clean(show_online_raw) and _parse_bool(show_online_raw, None) is None:
            errors.append({"row": row_num, "field": "show_online", "message": "Invalid boolean"})
        item_type = _clean(row.get("item_type")).upper()
        if item_type and item_type not in ("PRODUCT", "SERVICE"):
            errors.append({"row": row_num, "field": "item_type", "message": "Item type must be PRODUCT or SERVICE"})
        price_tax_mode = _clean(row.get("price_tax_mode")).upper()
        if price_tax_mode and price_tax_mode not in ("WITH_TAX", "WITHOUT_TAX"):
            errors.append(
                {"row": row_num, "field": "price_tax_mode", "message": "Price tax mode must be WITH_TAX or WITHOUT_TAX"}
            )
        expiry_date = row.get("expiry_date")
        if _clean(expiry_date) and _clean_date_string(expiry_date) is None:
            errors.append({"row": row_num, "field": "expiry_date", "message": "Invalid expiry date"})

    if errors:
        return {"status": "FAILED", "imported": 0, "errors": errors}

    for idx, row in enumerate(rows):
        row_num = idx + 2
        product = _resolve_product(
            db,
            user,
            row.get("product_code"),
            row.get("name"),
            row,
            errors,
            row_num,
        )
        if product is None:
            continue
        product.name = row.get("name", "").strip() or product.name
        if row.get("product_code", "").strip():
            product.product_code = row.get("product_code", "").strip()
        _upsert_stock_qty(
            db,
            user,
            product,
            _parse_float(row.get("opening_stock")),
            notes="Product import opening stock sync",
        )

    if errors:
        db.rollback()
        return {"status": "FAILED", "imported": 0, "errors": errors}

    db.commit()
    return {"status": "SUCCESS", "imported": len(rows), "errors": None}


def import_inventory_report(
    db: Session,
    user: User,
    rows: list[dict[str, str]],
    headers: list[str],
    *,
    options: dict[str, Any] | None = None,
) -> dict:
    required = {"product_name", "qty_on_hand"}
    if not required.issubset(set(headers)):
        missing = ", ".join(sorted(required - set(headers)))
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Missing required columns: {missing}")

    options = {
        "create_missing_products": True,
        "create_missing_categories": True,
        "update_existing_prices": False,
        "update_existing_reorder_level": True,
        "update_stock": True,
        "fallback_category_name": "Imported Inventory",
        **(options or {}),
    }
    errors: list[dict] = []
    stats = {
        "total_rows": len(rows),
        "products_created": 0,
        "categories_created": 0,
        "matched_existing_products": 0,
        "stock_adjustments_applied": 0,
        "rows_skipped": 0,
    }
    name_counts: dict[str, int] = {}

    for idx, row in enumerate(rows):
        row_num = idx + 2
        product_name = " ".join(_clean(row.get("product_name")).split())
        if not product_name:
            errors.append({"row": row_num, "field": "item_name", "message": "Blank item name skipped"})
            stats["rows_skipped"] += 1
            continue
        normalized_name = product_name.lower()
        name_counts[normalized_name] = name_counts.get(normalized_name, 0) + 1

        qty = _parse_float(row.get("qty_on_hand"))
        if qty is None:
            errors.append({"row": row_num, "field": "current_stock", "message": "Invalid current stock"})
            stats["rows_skipped"] += 1
            continue

        for source_field, field_label in (
            ("selling_rate", "sale_price"),
            ("purchase_rate", "purchase_price"),
            ("reorder_level", "minimum_stock"),
        ):
            raw = row.get(source_field)
            if _clean(raw) and _parse_float(raw) is None:
                errors.append({"row": row_num, "field": field_label, "message": "Invalid number ignored"})

    imported = 0
    for idx, row in enumerate(rows):
        row_num = idx + 2
        product_name = " ".join(_clean(row.get("product_name")).split())
        if not product_name:
            continue
        normalized_name = product_name.lower()
        if name_counts.get(normalized_name, 0) > 1:
            errors.append({"row": row_num, "field": "item_name", "message": "Duplicate item name in file skipped"})
            stats["rows_skipped"] += 1
            continue
        qty = _parse_float(row.get("qty_on_hand"))
        if qty is None:
            continue

        category_name = _clean(row.get("category"))
        if not category_name and options.get("create_missing_categories", True):
            category_name = options.get("fallback_category_name", "Imported Inventory")

        product = _resolve_product(
            db,
            user,
            None,
            product_name,
            {
                "category": category_name,
                "selling_rate": row.get("selling_rate"),
                "purchase_rate": row.get("purchase_rate"),
                "reorder_level": row.get("reorder_level"),
            },
            errors,
            row_num,
            options=options,
            stats=stats,
        )
        if product is None:
            stats["rows_skipped"] += 1
            continue

        if options.get("update_stock", True):
            previous_stock = (
                db.query(StockItem)
                .filter(StockItem.company_id == user.company_id, StockItem.product_id == product.id)
                .first()
            )
            previous_qty = float(previous_stock.qty_on_hand) if previous_stock and previous_stock.qty_on_hand is not None else 0.0
            _upsert_stock_qty(
                db,
                user,
                product,
                qty,
                notes="Imported from Stock Summary Report XLSX",
            )
            if previous_qty != float(qty):
                stats["stock_adjustments_applied"] += 1
        imported += 1

    db.commit()
    status = "SUCCESS"
    if errors and imported > 0:
        status = "PARTIAL_SUCCESS"
    elif errors and imported == 0:
        status = "FAILED"

    return {
        "status": status,
        "imported": imported,
        "errors": errors or None,
        "summary": stats,
    }


def import_inventory(db: Session, user: User, rows: list[dict[str, str]], headers: list[str]) -> dict:
    required = {"product_name", "qty_on_hand", "category"}
    if not required.issubset(set(headers)):
        missing = ", ".join(sorted(required - set(headers)))
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Missing required columns: {missing}")

    errors: list[dict] = []
    seen_products: set[str] = set()

    for idx, row in enumerate(rows):
        row_num = idx + 2
        name = row.get("product_name", "").strip()
        code = row.get("product_code", "").strip()
        key = (code or name).lower()
        if not name and not code:
            errors.append({"row": row_num, "field": "product_name", "message": "Product name or code required"})
            continue
        category_name = row.get("category", "").strip()
        if not category_name:
            errors.append({"row": row_num, "field": "category", "message": "Category required"})
        if key in seen_products:
            errors.append({"row": row_num, "field": "product_code", "message": "Duplicate product row in file"})
        seen_products.add(key)
        qty = _parse_float(row.get("qty_on_hand"))
        if qty is None:
            errors.append({"row": row_num, "field": "qty_on_hand", "message": "Invalid quantity"})
        for field in ("selling_rate", "purchase_rate", "tax_rate"):
            raw = row.get(field)
            if _clean(raw) and _parse_float(raw) is None:
                errors.append({"row": row_num, "field": field, "message": "Invalid number"})
        taxable_raw = row.get("taxable")
        if _clean(taxable_raw) and _parse_bool(taxable_raw, None) is None:
            errors.append({"row": row_num, "field": "taxable", "message": "Invalid boolean"})

    if errors:
        return {"status": "FAILED", "imported": 0, "errors": errors}

    for idx, row in enumerate(rows):
        row_num = idx + 2
        product = _resolve_product(
            db,
            user,
            row.get("product_code"),
            row.get("product_name"),
            row,
            errors,
            row_num,
        )
        qty = _parse_float(row.get("qty_on_hand")) or 0
        if product is None:
            continue

        _upsert_stock_qty(
            db,
            user,
            product,
            qty,
            notes="Inventory import quantity sync",
        )

    if errors:
        db.rollback()
        return {"status": "FAILED", "imported": 0, "errors": errors}

    db.commit()
    return {"status": "SUCCESS", "imported": len(rows), "errors": None}


def import_purchases(db: Session, user: User, rows: list[dict[str, str]], headers: list[str]) -> dict:
    required = {"invoice_no", "invoice_date", "tax_mode", "supplier_name", "qty", "price"}
    if not required.issubset(set(headers)):
        missing = ", ".join(sorted(required - set(headers)))
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Missing required columns: {missing}")

    errors: list[dict] = []
    existing_products = db.query(Product).filter(Product.company_id == user.company_id).all()
    known_codes = {p.product_code.lower() for p in existing_products if p.product_code}
    known_names = {p.name.lower() for p in existing_products}
    groups: dict[str, dict[str, Any]] = {}

    for idx, row in enumerate(rows):
        row_num = idx + 2
        invoice_no = row.get("invoice_no", "").strip()
        if not invoice_no:
            errors.append({"row": row_num, "field": "invoice_no", "message": "Invoice number required"})
            continue
        invoice_date = _parse_date(row.get("invoice_date"))
        if not invoice_date:
            errors.append({"row": row_num, "field": "invoice_date", "message": "Invalid invoice date"})
        tax_mode_raw = row.get("tax_mode", "").strip().upper()
        if tax_mode_raw not in ("GST", "NON_GST"):
            errors.append({"row": row_num, "field": "tax_mode", "message": "Tax mode must be GST or NON_GST"})
        supplier_name = row.get("supplier_name", "").strip()
        if not supplier_name:
            errors.append({"row": row_num, "field": "supplier_name", "message": "Supplier name required"})
        qty = _parse_float(row.get("qty"))
        price = _parse_float(row.get("price"))
        if qty is None:
            errors.append({"row": row_num, "field": "qty", "message": "Invalid quantity"})
        if price is None:
            errors.append({"row": row_num, "field": "price", "message": "Invalid price"})
        discount_raw = row.get("discount_percent")
        if _clean(discount_raw):
            discount_val = _parse_float(discount_raw)
            if discount_val is None or discount_val < 0 or discount_val > 100:
                errors.append({"row": row_num, "field": "discount_percent", "message": "Invalid discount percent"})
        tax_rate_raw = row.get("tax_rate")
        if _clean(tax_rate_raw) and _parse_float(tax_rate_raw) is None:
            errors.append({"row": row_num, "field": "tax_rate", "message": "Invalid tax rate"})
        taxable_raw = row.get("taxable")
        if _clean(taxable_raw) and _parse_bool(taxable_raw, None) is None:
            errors.append({"row": row_num, "field": "taxable", "message": "Invalid boolean"})
        description = row.get("description") or row.get("product_name") or row.get("product_code")
        if not _clean(description):
            errors.append({"row": row_num, "field": "description", "message": "Description or product name required"})
        payment_mode = _clean(row.get("payment_mode"))
        if payment_mode:
            try:
                PaymentMode(payment_mode)
            except Exception:
                errors.append({"row": row_num, "field": "payment_mode", "message": "Invalid payment mode"})
        code_key = _clean(row.get("product_code")).lower()
        name_key = _clean(row.get("product_name")).lower()
        has_existing = (code_key and code_key in known_codes) or (name_key and name_key in known_names)
        category_name = _clean(row.get("category"))
        if not has_existing and not category_name:
            errors.append({"row": row_num, "field": "category", "message": "Category required for new product"})

        if invoice_no not in groups:
            groups[invoice_no] = {
                "meta": {
                    "invoice_no": invoice_no,
                    "invoice_date": invoice_date,
                    "tax_mode": tax_mode_raw,
                    "is_interstate": _parse_bool(row.get("is_interstate"), False),
                    "supplier_name": supplier_name,
                    "supplier_phone": row.get("supplier_phone") or None,
                    "supplier_gstin": row.get("supplier_gstin") or None,
                    "supplier_address": row.get("supplier_address") or None,
                    "payment_status": row.get("payment_status") or None,
                    "payment_mode": row.get("payment_mode") or None,
                    "payment_reference": row.get("payment_reference") or None,
                    "round_off": _parse_float(row.get("round_off")) or 0,
                },
                "lines": [],
            }
        else:
            meta = groups[invoice_no]["meta"]
            if meta["supplier_name"] != supplier_name:
                errors.append({"row": row_num, "field": "supplier_name", "message": "Supplier mismatch in invoice"})

        groups[invoice_no]["lines"].append((row, row_num))

    if errors:
        return {"status": "FAILED", "imported": 0, "errors": errors}

    # Check existing invoice numbers
    existing = (
        db.query(Invoice)
        .filter(Invoice.company_id == user.company_id, Invoice.invoice_type == InvoiceType.PURCHASE)
        .all()
    )
    existing_nos = {row.invoice_no for row in existing}
    for invoice_no in groups:
        if invoice_no in existing_nos:
            row_num = groups[invoice_no]["lines"][0][1]
            errors.append({"row": row_num, "field": "invoice_no", "message": f"Invoice {invoice_no} already exists"})

    if errors:
        return {"status": "FAILED", "imported": 0, "errors": errors}

    imported = 0
    for invoice_no, data in groups.items():
        meta = data["meta"]
        supplier = (
            db.query(Supplier)
            .filter(Supplier.company_id == user.company_id, func.lower(Supplier.name) == meta["supplier_name"].lower())
            .first()
        )
        if not supplier:
            supplier = Supplier(
                company_id=user.company_id,
                name=meta["supplier_name"],
                phone=meta["supplier_phone"],
                gstin=meta["supplier_gstin"],
                address=meta["supplier_address"],
            )
            db.add(supplier)
            db.flush()

        line_payloads = []
        for row, row_num in data["lines"]:
            product = _resolve_product(
                db,
                user,
                row.get("product_code"),
                row.get("product_name") or row.get("description"),
                row,
                errors,
                row_num,
            )
            description = row.get("description") or row.get("product_name") or ""
            qty = _parse_float(row.get("qty")) or 0
            price = _parse_float(row.get("price")) or 0
            discount = _parse_float(row.get("discount_percent")) or 0
            taxable = _parse_bool(row.get("taxable"), True)
            tax_rate = _parse_float(row.get("tax_rate"))
            line_payloads.append(
                {
                    "product_id": product.id if product else None,
                    "description": description,
                    "hsn": row.get("hsn") or (product.hsn if product else None),
                    "qty": qty,
                    "unit": row.get("unit") or (product.unit if product else None),
                    "price": price,
                    "discount_percent": discount,
                    "taxable": taxable if taxable is not None else True,
                    "tax_rate": tax_rate,
                }
            )

        payload = {
            "invoice_no": meta["invoice_no"],
            "invoice_date": meta["invoice_date"],
            "tax_mode": TaxMode(meta["tax_mode"]),
            "is_interstate": bool(meta["is_interstate"]),
            "supplier_id": supplier.id,
            "lines": line_payloads,
            "round_off": meta["round_off"],
            "payment_status": meta["payment_status"],
            "payment_mode": meta["payment_mode"],
            "payment_reference": meta["payment_reference"],
        }
        create_invoice(db, user, InvoiceType.PURCHASE, payload)
        imported += 1

    if errors:
        db.rollback()
        return {"status": "FAILED", "imported": 0, "errors": errors}

    return {"status": "SUCCESS", "imported": imported, "errors": None}


def import_sales(db: Session, user: User, rows: list[dict[str, str]], headers: list[str]) -> dict:
    required = {"invoice_no", "invoice_date", "tax_mode", "customer_name", "qty", "price"}
    if not required.issubset(set(headers)):
        missing = ", ".join(sorted(required - set(headers)))
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Missing required columns: {missing}")

    errors: list[dict] = []
    existing_products = db.query(Product).filter(Product.company_id == user.company_id).all()
    known_codes = {p.product_code.lower() for p in existing_products if p.product_code}
    known_names = {p.name.lower() for p in existing_products}
    groups: dict[str, dict[str, Any]] = {}

    for idx, row in enumerate(rows):
        row_num = idx + 2
        invoice_no = row.get("invoice_no", "").strip()
        if not invoice_no:
            errors.append({"row": row_num, "field": "invoice_no", "message": "Invoice number required"})
            continue
        invoice_date = _parse_date(row.get("invoice_date"))
        if not invoice_date:
            errors.append({"row": row_num, "field": "invoice_date", "message": "Invalid invoice date"})
        tax_mode_raw = row.get("tax_mode", "").strip().upper()
        if tax_mode_raw not in ("GST", "NON_GST"):
            errors.append({"row": row_num, "field": "tax_mode", "message": "Tax mode must be GST or NON_GST"})
        customer_name = row.get("customer_name", "").strip()
        if not customer_name:
            errors.append({"row": row_num, "field": "customer_name", "message": "Customer name required"})
        qty = _parse_float(row.get("qty"))
        price = _parse_float(row.get("price"))
        if qty is None:
            errors.append({"row": row_num, "field": "qty", "message": "Invalid quantity"})
        if price is None:
            errors.append({"row": row_num, "field": "price", "message": "Invalid price"})
        discount_raw = row.get("discount_percent")
        if _clean(discount_raw):
            discount_val = _parse_float(discount_raw)
            if discount_val is None or discount_val < 0 or discount_val > 100:
                errors.append({"row": row_num, "field": "discount_percent", "message": "Invalid discount percent"})
        tax_rate_raw = row.get("tax_rate")
        if _clean(tax_rate_raw) and _parse_float(tax_rate_raw) is None:
            errors.append({"row": row_num, "field": "tax_rate", "message": "Invalid tax rate"})
        taxable_raw = row.get("taxable")
        if _clean(taxable_raw) and _parse_bool(taxable_raw, None) is None:
            errors.append({"row": row_num, "field": "taxable", "message": "Invalid boolean"})
        description = row.get("description") or row.get("product_name") or row.get("product_code")
        if not _clean(description):
            errors.append({"row": row_num, "field": "description", "message": "Description or product name required"})
        payment_mode = _clean(row.get("payment_mode"))
        if payment_mode:
            try:
                PaymentMode(payment_mode)
            except Exception:
                errors.append({"row": row_num, "field": "payment_mode", "message": "Invalid payment mode"})
        if tax_mode_raw == "GST":
            gstin = _clean(row.get("customer_gstin"))
            if not gstin:
                errors.append({"row": row_num, "field": "customer_gstin", "message": "GSTIN required for GST sales"})
        code_key = _clean(row.get("product_code")).lower()
        name_key = _clean(row.get("product_name")).lower()
        has_existing = (code_key and code_key in known_codes) or (name_key and name_key in known_names)
        category_name = _clean(row.get("category"))
        if not has_existing and not category_name:
            errors.append({"row": row_num, "field": "category", "message": "Category required for new product"})

        if invoice_no not in groups:
            groups[invoice_no] = {
                "meta": {
                    "invoice_no": invoice_no,
                    "invoice_date": invoice_date,
                    "tax_mode": tax_mode_raw,
                    "is_interstate": _parse_bool(row.get("is_interstate"), False),
                    "customer_name": customer_name,
                    "customer_phone": row.get("customer_phone") or None,
                    "customer_gstin": row.get("customer_gstin") or None,
                    "customer_address": row.get("customer_address") or None,
                    "payment_mode": row.get("payment_mode") or None,
                    "payment_reference": row.get("payment_reference") or None,
                    "round_off": _parse_float(row.get("round_off")) or 0,
                },
                "lines": [],
            }
        else:
            meta = groups[invoice_no]["meta"]
            if meta["customer_name"] != customer_name:
                errors.append({"row": row_num, "field": "customer_name", "message": "Customer mismatch in invoice"})

        groups[invoice_no]["lines"].append((row, row_num))

    if errors:
        return {"status": "FAILED", "imported": 0, "errors": errors}

    existing = (
        db.query(Invoice)
        .filter(Invoice.company_id == user.company_id, Invoice.invoice_type == InvoiceType.SALES)
        .all()
    )
    existing_nos = {row.invoice_no for row in existing}
    for invoice_no in groups:
        if invoice_no in existing_nos:
            row_num = groups[invoice_no]["lines"][0][1]
            errors.append({"row": row_num, "field": "invoice_no", "message": f"Invoice {invoice_no} already exists"})

    if errors:
        return {"status": "FAILED", "imported": 0, "errors": errors}

    imported = 0
    for invoice_no, data in groups.items():
        meta = data["meta"]
        customer = (
            db.query(Customer)
            .filter(Customer.company_id == user.company_id, func.lower(Customer.name) == meta["customer_name"].lower())
            .first()
        )
        if not customer:
            customer = Customer(
                company_id=user.company_id,
                name=meta["customer_name"],
                phone=meta["customer_phone"],
                gstin=meta["customer_gstin"],
                address=meta["customer_address"],
            )
            db.add(customer)
            db.flush()

        line_payloads = []
        for row, row_num in data["lines"]:
            product = _resolve_product(
                db,
                user,
                row.get("product_code"),
                row.get("product_name") or row.get("description"),
                row,
                errors,
                row_num,
            )
            description = row.get("description") or row.get("product_name") or ""
            qty = _parse_float(row.get("qty")) or 0
            price = _parse_float(row.get("price")) or 0
            discount = _parse_float(row.get("discount_percent")) or 0
            taxable = _parse_bool(row.get("taxable"), True)
            tax_rate = _parse_float(row.get("tax_rate"))
            line_payloads.append(
                {
                    "product_id": product.id if product else None,
                    "description": description,
                    "hsn": row.get("hsn") or (product.hsn if product else None),
                    "qty": qty,
                    "unit": row.get("unit") or (product.unit if product else None),
                    "price": price,
                    "discount_percent": discount,
                    "taxable": taxable if taxable is not None else True,
                    "tax_rate": tax_rate,
                }
            )

        payload = {
            "invoice_no": meta["invoice_no"],
            "invoice_date": meta["invoice_date"],
            "tax_mode": TaxMode(meta["tax_mode"]),
            "is_interstate": bool(meta["is_interstate"]),
            "customer_id": customer.id,
            "lines": line_payloads,
            "round_off": meta["round_off"],
            "payment_mode": meta["payment_mode"],
            "payment_reference": meta["payment_reference"],
        }
        create_invoice(db, user, InvoiceType.SALES, payload)
        imported += 1

    if errors:
        db.rollback()
        return {"status": "FAILED", "imported": 0, "errors": errors}

    return {"status": "SUCCESS", "imported": imported, "errors": None}
