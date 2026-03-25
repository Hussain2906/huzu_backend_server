import csv
import io

import pytest
from openpyxl import Workbook

from testing.tests.factories import create_company, create_user
from testing.tests.helpers import auth_headers, login


PARTY_HEADERS = [
    "SR NO",
    "NAME",
    "PHONE",
    "CATEGORY",
    "CREDIT",
    "TYPE",
    "GST NO",
    "BILLING TYPE",
    "DOB",
    "BUSINESS NAME",
    "EMAIL",
    "BILLING ADDRESS",
    "BILLING STATES & U.T.",
    "BILLING POSTAL CODE",
    "DELIVERY ADDRESS",
    "DELIVERY STATES & U.T.",
    "DELIVERY POSTAL CODE",
    "PAYMENT TERM",
    "SEND ALERTS",
    "FAVOURITE PARTY",
]



def _csv_bytes(rows: list[dict[str, str]]) -> bytes:
    stream = io.StringIO()
    writer = csv.DictWriter(stream, fieldnames=PARTY_HEADERS)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return stream.getvalue().encode("utf-8")



def _xlsx_bytes_with_title_row(rows: list[dict[str, str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Party Details Report"
    for index, header in enumerate(PARTY_HEADERS, start=1):
        sheet.cell(row=3, column=index, value=header)
    for row_index, row in enumerate(rows, start=4):
        for col_index, header in enumerate(PARTY_HEADERS, start=1):
            sheet.cell(row=row_index, column=col_index, value=row.get(header, ""))

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


@pytest.mark.integration
@pytest.mark.csv
def test_party_import_preview_and_commit_create_customer(client, db_session, seed_roles):
    company = create_company(db_session, name="Party Import Co")
    create_user(db_session, username="owner", password="Pass@1234", company_id=company.id, role_code="SUPER_ADMIN")
    access_token, _ = login(client, "owner", "Pass@1234")

    csv_data = _csv_bytes(
        [
            {
                "SR NO": "1",
                "NAME": "Arihant Traders",
                "PHONE": "9000000000",
                "CATEGORY": "DEBTORS",
                "CREDIT": "₹ 24005.0",
                "TYPE": "CUSTOMER",
                "GST NO": "22AAAAA0000A1Z5",
                "BILLING TYPE": "REGULAR",
                "DOB": "",
                "BUSINESS NAME": "Arihant",
                "EMAIL": "sales@arihant.example",
                "BILLING ADDRESS": "Main Street",
                "BILLING STATES & U.T.": "Maharashtra",
                "BILLING POSTAL CODE": "411001",
                "DELIVERY ADDRESS": "Warehouse Road",
                "DELIVERY STATES & U.T.": "Maharashtra",
                "DELIVERY POSTAL CODE": "411002",
                "PAYMENT TERM": "30 DAYS",
                "SEND ALERTS": "YES",
                "FAVOURITE PARTY": "NO",
            }
        ]
    )

    preview = client.post(
        "/v1/imports/parties/preview",
        files={"file": ("party.csv", csv_data, "text/csv")},
        headers=auth_headers(access_token),
    )
    assert preview.status_code == 200
    body = preview.json()
    assert body["summary"]["create"] == 1
    assert body["rows"][0]["action"] == "CREATE"

    commit = client.post(
        "/v1/imports/parties/commit",
        json={"batch_id": body["batch_id"], "duplicate_policy": "UPDATE_MATCHED"},
        headers=auth_headers(access_token),
    )
    assert commit.status_code == 200
    commit_body = commit.json()
    assert commit_body["status"] == "COMPLETED"
    assert commit_body["summary"]["created"] == 1

    customers = client.get("/v1/masters/customers", headers=auth_headers(access_token))
    assert customers.status_code == 200
    assert len(customers.json()) == 1
    assert customers.json()[0]["party_id"]

    parties = client.get("/v1/masters/parties?role=CUSTOMER", headers=auth_headers(access_token))
    assert parties.status_code == 200
    assert len(parties.json()) == 1


@pytest.mark.integration
@pytest.mark.csv
def test_party_import_updates_existing_party_and_adds_second_role(client, db_session, seed_roles):
    company = create_company(db_session, name="Party Import Co")
    create_user(db_session, username="owner", password="Pass@1234", company_id=company.id, role_code="SUPER_ADMIN")
    access_token, _ = login(client, "owner", "Pass@1234")

    existing_customer = client.post(
        "/v1/masters/customers",
        json={"name": "Zen Mart", "phone": "9000000001", "gstin": "27AAACB1234C1Z9", "address": "Old Address"},
        headers=auth_headers(access_token),
    )
    assert existing_customer.status_code == 200
    party_id = existing_customer.json()["party_id"]

    csv_data = _csv_bytes(
        [
            {
                "SR NO": "2",
                "NAME": "Zen Mart Pvt",
                "PHONE": "9000000001",
                "CATEGORY": "CREDITORS",
                "CREDIT": "₹ -867.0",
                "TYPE": "SUPPLIER",
                "GST NO": "27AAACB1234C1Z9",
                "BILLING TYPE": "REGULAR",
                "DOB": "",
                "BUSINESS NAME": "Zen Mart Pvt Ltd",
                "EMAIL": "procurement@zen.example",
                "BILLING ADDRESS": "Market Road",
                "BILLING STATES & U.T.": "Maharashtra",
                "BILLING POSTAL CODE": "411010",
                "DELIVERY ADDRESS": "",
                "DELIVERY STATES & U.T.": "",
                "DELIVERY POSTAL CODE": "",
                "PAYMENT TERM": "15 DAYS",
                "SEND ALERTS": "NO",
                "FAVOURITE PARTY": "YES",
            }
        ]
    )

    preview = client.post(
        "/v1/imports/parties/preview",
        files={"file": ("party.csv", csv_data, "text/csv")},
        headers=auth_headers(access_token),
    )
    assert preview.status_code == 200
    body = preview.json()
    assert body["rows"][0]["action"] == "UPDATE"
    assert body["rows"][0]["matched_party_id"] == party_id

    commit = client.post(
        "/v1/imports/parties/commit",
        json={"batch_id": body["batch_id"], "duplicate_policy": "UPDATE_MATCHED"},
        headers=auth_headers(access_token),
    )
    assert commit.status_code == 200
    assert commit.json()["status"] == "COMPLETED"
    assert commit.json()["summary"]["updated"] == 1

    suppliers = client.get("/v1/masters/suppliers", headers=auth_headers(access_token))
    assert suppliers.status_code == 200
    assert len(suppliers.json()) == 1
    assert suppliers.json()[0]["party_id"] == party_id

    party_detail = client.get(f"/v1/masters/parties/{party_id}", headers=auth_headers(access_token))
    assert party_detail.status_code == 200
    assert sorted(party_detail.json()["roles"]) == ["CUSTOMER", "SUPPLIER"]


@pytest.mark.integration
@pytest.mark.csv
def test_party_import_preview_marks_invalid_credit_as_error(client, db_session, seed_roles):
    company = create_company(db_session, name="Party Import Co")
    create_user(db_session, username="owner", password="Pass@1234", company_id=company.id, role_code="SUPER_ADMIN")
    access_token, _ = login(client, "owner", "Pass@1234")

    csv_data = _csv_bytes(
        [
            {
                "SR NO": "3",
                "NAME": "Bad Credit Co",
                "PHONE": "9000000009",
                "CATEGORY": "DEBTORS",
                "CREDIT": "₹ ABC",
                "TYPE": "CUSTOMER",
                "GST NO": "",
                "BILLING TYPE": "REGULAR",
                "DOB": "",
                "BUSINESS NAME": "",
                "EMAIL": "",
                "BILLING ADDRESS": "",
                "BILLING STATES & U.T.": "",
                "BILLING POSTAL CODE": "",
                "DELIVERY ADDRESS": "",
                "DELIVERY STATES & U.T.": "",
                "DELIVERY POSTAL CODE": "",
                "PAYMENT TERM": "",
                "SEND ALERTS": "NO",
                "FAVOURITE PARTY": "NO",
            }
        ]
    )

    preview = client.post(
        "/v1/imports/parties/preview",
        files={"file": ("party.csv", csv_data, "text/csv")},
        headers=auth_headers(access_token),
    )
    assert preview.status_code == 200
    row = preview.json()["rows"][0]
    assert row["action"] == "ERROR"
    assert any(err["field"] == "CREDIT" for err in row["errors"])


@pytest.mark.integration
def test_party_import_preview_accepts_xlsx_with_title_rows(client, db_session, seed_roles):
    company = create_company(db_session, name="Party Import Co")
    create_user(db_session, username="owner", password="Pass@1234", company_id=company.id, role_code="SUPER_ADMIN")
    access_token, _ = login(client, "owner", "Pass@1234")

    xlsx_data = _xlsx_bytes_with_title_row(
        [
            {
                "SR NO": "1",
                "NAME": "Title Row Works",
                "PHONE": "9000000010",
                "CATEGORY": "DEBTORS",
                "CREDIT": "₹ 120.0",
                "TYPE": "CUSTOMER",
                "GST NO": "",
                "BILLING TYPE": "REGULAR",
                "DOB": "",
                "BUSINESS NAME": "",
                "EMAIL": "",
                "BILLING ADDRESS": "Sample Address",
                "BILLING STATES & U.T.": "Maharashtra",
                "BILLING POSTAL CODE": "411001",
                "DELIVERY ADDRESS": "",
                "DELIVERY STATES & U.T.": "",
                "DELIVERY POSTAL CODE": "",
                "PAYMENT TERM": "30 DAYS",
                "SEND ALERTS": "YES",
                "FAVOURITE PARTY": "NO",
            }
        ]
    )

    preview = client.post(
        "/v1/imports/parties/preview",
        files={"file": ("party.xlsx", xlsx_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers(access_token),
    )
    assert preview.status_code == 200
    body = preview.json()
    assert body["summary"]["create"] == 1
    assert body["rows"][0]["action"] == "CREATE"
    assert body["rows"][0]["row_number"] == 4
